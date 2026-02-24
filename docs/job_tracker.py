"""
DevOps Job Radar — Multi-Source Edition
Sources: LinkedIn (RSS), Indeed (RSS), Adzuna (API), Arbeitnow (API),
         RemoteOK (API), WorkingNomads (RSS), TheMuse (API)
Scores jobs against Seetharam's resume → Excel + docs/jobs.json for GitHub Pages.
"""

import os, re, json, smtplib, logging, hashlib, time, random
import requests
import feedparser
from bs4 import BeautifulSoup
from datetime import datetime, timedelta, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from pathlib import Path
from urllib.parse import quote_plus
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

# ─── Config ────────────────────────────────────────────────────────────────────
MATCH_THRESHOLD = 55
MAX_AGE_HOURS   = 24
DOCS_DIR        = "docs"

HTTP_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9,de;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

SEARCH_TERMS = [
    "DevOps Engineer",
    "Site Reliability Engineer",
    "Platform Engineer",
    "Cloud Infrastructure Engineer",
    "Kubernetes Engineer",
    "SRE Engineer",
]

# ─── Resume Profile ────────────────────────────────────────────────────────────
RESUME_SKILLS = {
    "kubernetes": ["kubernetes","k8s","eks","aks","helm","kustomize","kubectl","karpenter"],
    "cicd":       ["ci/cd","cicd","github actions","gitlab ci","azure devops","jenkins","pipeline","gitops"],
    "cloud":      ["azure","aws","gcp","cloud","eks","aks","ec2","s3"],
    "iac":        ["terraform","infrastructure as code","iac","ansible","pulumi","cloudformation"],
    "containers": ["docker","containers","containerization","helm","podman"],
    "monitoring": ["prometheus","grafana","loki","observability","monitoring","fluent-bit","datadog","elk"],
    "security":   ["sonarqube","trivy","zero trust","waf","cert-manager","owasp","sast","vault"],
    "scripting":  ["python","bash","shell","scripting","go","golang"],
    "gitops":     ["argocd","gitops","argo rollouts","flux","blue-green","canary","progressive delivery"],
    "platform":   ["platform engineering","sre","site reliability","devops","infrastructure"],
}

SKILL_WEIGHTS = {
    "kubernetes":20,"cicd":18,"cloud":17,"iac":15,
    "containers":12,"gitops":10,"monitoring":8,
    "security":7,"platform":7,"scripting":6,
}

# ─── Helpers ───────────────────────────────────────────────────────────────────
def _clean_html(text: str) -> str:
    return BeautifulSoup(text, "html.parser").get_text(" ")

def _uid(url: str) -> str:
    return hashlib.md5(url.encode()).hexdigest()

def _extract_location(text: str) -> str:
    cities = ["Berlin","Munich","München","Hamburg","Frankfurt","Cologne","Köln",
              "Stuttgart","Düsseldorf","Krefeld","Dortmund","Leipzig","Dresden",
              "Nuremberg","Nürnberg","Hannover","Bremen","Heidelberg","Remote"]
    for city in cities:
        if city.lower() in text.lower():
            return f"{city}, Germany"
    return "Germany"

def classify_remote_text(text: str) -> str:
    t = text.lower()
    if any(x in t for x in ["fully remote","100% remote","full remote","fully-remote"]):
        return "Fully Remote"
    if any(x in t for x in ["remote","home office","homeoffice","distributed"]):
        return "Hybrid / Remote"
    return "On-site"

def _parse_date(entry) -> tuple[datetime | None, str]:
    pub = entry.get("published_parsed")
    if pub:
        dt = datetime(*pub[:6], tzinfo=timezone.utc)
        return dt, dt.strftime("%d %b %Y %H:%M")
    return None, datetime.now().strftime("%d %b %Y")

def _too_old(dt, hours=MAX_AGE_HOURS) -> bool:
    if dt is None:
        return False
    return datetime.now(timezone.utc) - dt > timedelta(hours=hours)

# ═══════════════════════════════════════════════════════════════════════════════
# SOURCE 1 — LINKEDIN RSS
# ═══════════════════════════════════════════════════════════════════════════════
def fetch_linkedin(query: str) -> list[dict]:
    encoded = quote_plus(query)
    rss_url = (
        f"https://www.linkedin.com/jobs/search/?keywords={encoded}"
        f"&location=Germany&f_TPR=r86400&f_WT=2&format=rss"
    )
    results = []
    try:
        feed = feedparser.parse(rss_url)
        for entry in feed.entries:
            dt, date_str = _parse_date(entry)
            if _too_old(dt):
                continue
            desc = _clean_html(entry.get("summary",""))
            title = entry.get("title","").strip()
            company = title.split(" at ")[-1].strip() if " at " in title else \
                      title.split(" bei ")[-1].strip() if " bei " in title else ""
            results.append({
                "_id": _uid(entry.get("link","")),
                "_src": "LinkedIn",
                "title": title,
                "company": company,
                "location": _extract_location(desc),
                "description": desc,
                "date_posted": date_str,
                "apply_url": entry.get("link",""),
                "remote_type": classify_remote_text(title + " " + desc),
            })
    except Exception as e:
        log.warning(f"LinkedIn RSS [{query}]: {e}")
    return results


# ═══════════════════════════════════════════════════════════════════════════════
# SOURCE 2 — INDEED RSS
# ═══════════════════════════════════════════════════════════════════════════════
def fetch_indeed(query: str) -> list[dict]:
    encoded = quote_plus(query)
    rss_url = (
        f"https://de.indeed.com/rss?q={encoded}&l=Germany"
        f"&fromage=1&remotejob=032b3046-06a3-4876-8dfd-474eb5e7ed11"
    )
    results = []
    try:
        feed = feedparser.parse(rss_url)
        for entry in feed.entries:
            dt, date_str = _parse_date(entry)
            if _too_old(dt, hours=MAX_AGE_HOURS + 12):
                continue
            desc = _clean_html(entry.get("summary",""))
            results.append({
                "_id": _uid(entry.get("link","")),
                "_src": "Indeed",
                "title": entry.get("title","").strip(),
                "company": entry.get("author","") or "",
                "location": _extract_location(desc),
                "description": desc,
                "date_posted": date_str,
                "apply_url": entry.get("link",""),
                "remote_type": classify_remote_text(entry.get("title","") + " " + desc),
            })
    except Exception as e:
        log.warning(f"Indeed RSS [{query}]: {e}")
    return results


# ═══════════════════════════════════════════════════════════════════════════════
# SOURCE 3 — ADZUNA (free API — 250 req/day)
# ═══════════════════════════════════════════════════════════════════════════════
def fetch_adzuna(query: str) -> list[dict]:
    app_id  = os.getenv("ADZUNA_APP_ID")
    api_key = os.getenv("ADZUNA_API_KEY")
    if not app_id or not api_key:
        return []
    try:
        r = requests.get(
            "https://api.adzuna.com/v1/api/jobs/de/search/1",
            params={"app_id":app_id,"app_key":api_key,"results_per_page":20,
                    "what":query,"max_days_old":1,"content-type":"application/json"},
            headers=HTTP_HEADERS, timeout=15)
        r.raise_for_status()
        results = []
        for job in r.json().get("results",[]):
            created = job.get("created","")
            try:
                dt = datetime.fromisoformat(created.replace("Z","+00:00"))
                date_str = dt.strftime("%d %b %Y %H:%M")
            except Exception:
                date_str = datetime.now().strftime("%d %b %Y")
            desc = re.sub(r"<[^>]+>"," ", job.get("description",""))
            results.append({
                "_id": str(job.get("id","")),
                "_src": "Adzuna",
                "title": job.get("title",""),
                "company": job.get("company",{}).get("display_name",""),
                "location": job.get("location",{}).get("display_name","Germany"),
                "description": desc,
                "date_posted": date_str,
                "apply_url": job.get("redirect_url",""),
                "remote_type": classify_remote_text(job.get("title","") + " " + desc),
            })
        return results
    except Exception as e:
        log.warning(f"Adzuna [{query}]: {e}")
        return []


# ═══════════════════════════════════════════════════════════════════════════════
# SOURCE 4 — ARBEITNOW (free API — Germany-focused)
# ═══════════════════════════════════════════════════════════════════════════════
def fetch_arbeitnow() -> list[dict]:
    tags = ["devops","kubernetes","terraform","sre","platform-engineering","cloud","ansible"]
    results, seen = [], set()
    cutoff = datetime.now(timezone.utc) - timedelta(hours=MAX_AGE_HOURS)
    for tag in tags:
        try:
            r = requests.get(
                f"https://arbeitnow.com/api/job-board-api?tag={tag}&remote=true",
                headers=HTTP_HEADERS, timeout=15)
            r.raise_for_status()
            for job in r.json().get("data",[]):
                jid = str(job.get("slug",""))
                if jid in seen: continue
                seen.add(jid)
                try:
                    ca = job.get("created_at",0)
                    dt = datetime.fromtimestamp(ca,tz=timezone.utc) if isinstance(ca,int) \
                         else datetime.fromisoformat(str(ca).replace("Z","+00:00"))
                    if dt < cutoff: continue
                    date_str = dt.strftime("%d %b %Y %H:%M")
                except Exception:
                    date_str = datetime.now().strftime("%d %b %Y")
                desc = _clean_html(job.get("description",""))
                results.append({
                    "_id": jid,
                    "_src": "Arbeitnow",
                    "title": job.get("title",""),
                    "company": job.get("company_name",""),
                    "location": job.get("location","Germany") or "Germany",
                    "description": desc,
                    "date_posted": date_str,
                    "apply_url": job.get("url",""),
                    "remote_type": "Fully Remote" if job.get("remote") else "Hybrid / Remote",
                })
            time.sleep(0.4)
        except Exception as e:
            log.warning(f"Arbeitnow [{tag}]: {e}")
    return results


# ═══════════════════════════════════════════════════════════════════════════════
# SOURCE 5 — REMOTEOK (free public API)
# ═══════════════════════════════════════════════════════════════════════════════
def fetch_remoteok() -> list[dict]:
    cutoff = datetime.now(timezone.utc) - timedelta(hours=MAX_AGE_HOURS)
    relevant = {"devops","kubernetes","docker","terraform","aws","azure","gcp",
                "sre","cloud","platform","infrastructure","ci-cd","ansible","helm"}
    try:
        r = requests.get("https://remoteok.com/api",
                         headers={**HTTP_HEADERS,"Accept":"application/json"}, timeout=20)
        r.raise_for_status()
        results = []
        for job in r.json()[1:]:
            if not isinstance(job, dict): continue
            tags = {t.lower() for t in job.get("tags",[])}
            if not relevant.intersection(tags): continue
            loc = job.get("location","").lower()
            if loc and not any(x in loc for x in ["germany","europe","eu","worldwide",""]):
                continue
            epoch = job.get("epoch",0)
            if epoch:
                dt = datetime.fromtimestamp(epoch, tz=timezone.utc)
                if dt < cutoff: continue
                date_str = dt.strftime("%d %b %Y %H:%M")
            else:
                date_str = datetime.now().strftime("%d %b %Y")
            desc = _clean_html(job.get("description",""))
            results.append({
                "_id": str(job.get("id","")),
                "_src": "RemoteOK",
                "title": job.get("position",""),
                "company": job.get("company",""),
                "location": job.get("location","Remote") or "Remote (Worldwide)",
                "description": desc,
                "date_posted": date_str,
                "apply_url": job.get("url","") or f"https://remoteok.com/remote-jobs/{job.get('id','')}",
                "remote_type": "Fully Remote",
            })
        return results
    except Exception as e:
        log.warning(f"RemoteOK: {e}")
        return []


# ═══════════════════════════════════════════════════════════════════════════════
# SOURCE 6 — WORKINGNOMADS RSS
# ═══════════════════════════════════════════════════════════════════════════════
def fetch_workingnomads() -> list[dict]:
    feeds = [
        "https://www.workingnomads.com/feed?category=devops",
        "https://www.workingnomads.com/feed?category=sysadmin",
        "https://www.workingnomads.com/feed?category=cloud",
    ]
    results, seen = [], set()
    cutoff = datetime.now(timezone.utc) - timedelta(hours=MAX_AGE_HOURS * 2)
    for rss_url in feeds:
        try:
            feed = feedparser.parse(rss_url)
            for entry in feed.entries:
                link = entry.get("link","")
                if link in seen: continue
                seen.add(link)
                dt, date_str = _parse_date(entry)
                if _too_old(dt, hours=MAX_AGE_HOURS * 2): continue
                desc = _clean_html(entry.get("summary",""))
                results.append({
                    "_id": _uid(link),
                    "_src": "WorkingNomads",
                    "title": entry.get("title",""),
                    "company": entry.get("author",""),
                    "location": "Remote (Worldwide)",
                    "description": desc,
                    "date_posted": date_str,
                    "apply_url": link,
                    "remote_type": "Fully Remote",
                })
        except Exception as e:
            log.warning(f"WorkingNomads: {e}")
    return results


# ═══════════════════════════════════════════════════════════════════════════════
# SOURCE 7 — THE MUSE (free API)
# ═══════════════════════════════════════════════════════════════════════════════
def fetch_themuse() -> list[dict]:
    categories = ["Software Engineer","DevOps","Infrastructure","IT"]
    results = []
    cutoff = datetime.now(timezone.utc) - timedelta(hours=MAX_AGE_HOURS * 3)
    for cat in categories:
        try:
            r = requests.get(
                "https://www.themuse.com/api/public/jobs",
                params={"category":cat,"location":"Remote","page":1,"descending":"true"},
                headers=HTTP_HEADERS, timeout=15)
            r.raise_for_status()
            for job in r.json().get("results",[]):
                pub = job.get("publication_date","")
                try:
                    dt = datetime.fromisoformat(pub.replace("Z","+00:00"))
                    if dt < cutoff: continue
                    date_str = dt.strftime("%d %b %Y %H:%M")
                except Exception:
                    date_str = datetime.now().strftime("%d %b %Y")
                desc = " ".join(_clean_html(s.get("body","")) for s in job.get("contents",[]))
                locs = [loc.get("name","") for loc in job.get("locations",[])]
                results.append({
                    "_id": str(job.get("id","")),
                    "_src": "TheMuse",
                    "title": job.get("name",""),
                    "company": job.get("company",{}).get("name",""),
                    "location": ", ".join(locs) or "Remote",
                    "description": desc,
                    "date_posted": date_str,
                    "apply_url": job.get("refs",{}).get("landing_page",""),
                    "remote_type": "Fully Remote",
                })
            time.sleep(0.4)
        except Exception as e:
            log.warning(f"TheMuse [{cat}]: {e}")
    return results


# ═══════════════════════════════════════════════════════════════════════════════
# SCORING
# ═══════════════════════════════════════════════════════════════════════════════
def score_job(job: dict) -> tuple[int, list[str]]:
    text = (job.get("title","") + " " + job.get("description","")).lower()
    total_weight = sum(SKILL_WEIGHTS.values())
    earned, matched = 0, []
    for category, keywords in RESUME_SKILLS.items():
        for kw in keywords:
            if kw in text:
                earned += SKILL_WEIGHTS[category]
                matched.append(kw.title())
                break
    if any(x in text for x in ["remote","home office","homeoffice","distributed"]):
        earned += 5
    loc = job.get("location","").lower()
    if any(x in loc for x in ["germany","deutschland","berlin","munich","hamburg","düsseldorf"]):
        earned += 3
    return min(100, int((earned / total_weight) * 100)), list(dict.fromkeys(matched))


def get_career_url(job: dict) -> str:
    url = job.get("apply_url","")
    if url: return url
    company = job.get("company","")
    if company:
        slug = re.sub(r"[^a-z0-9]","", company.lower())
        return f"https://www.{slug}.com/careers"
    return "#"


def summarize(text: str, length: int = 240) -> str:
    clean = re.sub(r"\s+"," ", text).strip()
    return (clean[:length] + "…") if len(clean) > length else clean


# ═══════════════════════════════════════════════════════════════════════════════
# ORCHESTRATOR
# ═══════════════════════════════════════════════════════════════════════════════
def collect_all_jobs() -> list[dict]:
    raw: list[dict] = []

    log.info("── Source 1/7: LinkedIn RSS ──")
    for q in SEARCH_TERMS:
        raw.extend(fetch_linkedin(q))
        time.sleep(random.uniform(0.8, 1.5))

    log.info("── Source 2/7: Indeed RSS ──")
    for q in SEARCH_TERMS:
        raw.extend(fetch_indeed(q))
        time.sleep(random.uniform(0.8, 1.5))

    log.info("── Source 3/7: Adzuna API ──")
    for q in SEARCH_TERMS:
        raw.extend(fetch_adzuna(q))
        time.sleep(0.3)

    log.info("── Source 4/7: Arbeitnow API ──")
    raw.extend(fetch_arbeitnow())

    log.info("── Source 5/7: RemoteOK API ──")
    raw.extend(fetch_remoteok())

    log.info("── Source 6/7: WorkingNomads RSS ──")
    raw.extend(fetch_workingnomads())

    log.info("── Source 7/7: TheMuse API ──")
    raw.extend(fetch_themuse())

    # Deduplicate
    seen, unique = set(), []
    for job in raw:
        jid = job.get("_id","")
        if jid and jid not in seen:
            seen.add(jid)
            unique.append(job)

    log.info(f"Total unique jobs collected: {len(unique)}")
    return unique


def process_jobs(raw_jobs: list[dict]) -> list[dict]:
    processed = []
    for job in raw_jobs:
        score, matched = score_job(job)
        if score < MATCH_THRESHOLD:
            continue
        processed.append({
            "title":          job.get("title","N/A").strip(),
            "company":        (job.get("company","") or "Unknown").strip(),
            "location":       job.get("location","Germany"),
            "remote_type":    job.get("remote_type","Unknown"),
            "date_posted":    job.get("date_posted", datetime.now().strftime("%d %b %Y")),
            "score":          score,
            "matched_skills": matched,
            "summary":        summarize(job.get("description","")),
            "career_url":     get_career_url(job),
            "source":         job.get("_src","Unknown"),
        })
    processed.sort(key=lambda x: x["score"], reverse=True)
    log.info(f"Jobs above {MATCH_THRESHOLD}% threshold: {len(processed)}")
    return processed


# ═══════════════════════════════════════════════════════════════════════════════
# OUTPUTS
# ═══════════════════════════════════════════════════════════════════════════════
def export_json(jobs_data: list[dict], docs_dir: str = DOCS_DIR):
    Path(docs_dir).mkdir(parents=True, exist_ok=True)
    out = Path(docs_dir) / "jobs.json"
    out.write_text(json.dumps({
        "updated": datetime.utcnow().isoformat() + "Z",
        "count": len(jobs_data),
        "jobs": jobs_data,
    }, indent=2, ensure_ascii=False))
    log.info(f"jobs.json → {out}")


HEADERS_XL = ["Job Title","Company","Location","Remote Type","Source",
               "Date Posted","Match Score","Key Matching Skills","Role Summary","Apply URL"]
HEADER_COLOR = "1F3864"

def _border():
    s = Side(style="thin", color="B8CCE4")
    return Border(left=s, right=s, top=s, bottom=s)

def score_fill(score):
    color = "C6EFCE" if score >= 75 else "FFEB9C" if score >= 60 else "FCE4D6"
    return PatternFill("solid", fgColor=color)

def build_excel(jobs_data: list[dict], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Matching Jobs"
    ws.freeze_panes = "A3"

    ws.merge_cells(f"A1:{get_column_letter(len(HEADERS_XL))}1")
    tc = ws["A1"]
    tc.value = f"DevOps Job Matches — Germany Remote  |  {datetime.now().strftime('%d %b %Y %H:%M')}"
    tc.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    tc.fill = PatternFill("solid", fgColor=HEADER_COLOR)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    for col, h in enumerate(HEADERS_XL, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=HEADER_COLOR)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
    ws.row_dimensions[2].height = 22

    for row_idx, job in enumerate(jobs_data, start=3):
        fill = score_fill(job["score"])
        values = [job["title"],job["company"],job["location"],job["remote_type"],
                  job.get("source",""),job["date_posted"],job["score"],
                  ", ".join(job["matched_skills"]),job["summary"],job["career_url"]]
        for col_idx, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = Font(name="Arial", size=10)
            c.fill = fill
            c.border = _border()
            c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row_idx, column=7).font = Font(name="Arial", bold=True, size=11)
        ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal="center", vertical="top")
        url_c = ws.cell(row=row_idx, column=10)
        url_c.hyperlink = job["career_url"]
        url_c.font = Font(name="Arial", size=10, color="0563C1", underline="single")
        ws.row_dimensions[row_idx].height = 55

    for i, w in enumerate([36,22,20,18,14,16,12,36,55,40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS_XL))}{len(jobs_data)+2}"

    # Stats sheet
    ws2 = wb.create_sheet("Summary Stats")
    ws2.merge_cells("A1:B1")
    ws2["A1"].value = "Run Summary"
    ws2["A1"].font  = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws2["A1"].fill  = PatternFill("solid", fgColor=HEADER_COLOR)
    sources = {}
    for j in jobs_data:
        s = j.get("source","?")
        sources[s] = sources.get(s,0) + 1
    stats = [
        ("Total Matches", len(jobs_data)),
        ("High Match ≥75%", sum(1 for j in jobs_data if j["score"]>=75)),
        ("Good Match 60–74%", sum(1 for j in jobs_data if 60<=j["score"]<75)),
        ("Fully Remote", sum(1 for j in jobs_data if "Fully" in j["remote_type"])),
        ("Run Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("",""),("Jobs by Source",""),
        *[(f"  {src}", cnt) for src,cnt in sorted(sources.items())],
    ]
    for r,(label,val) in enumerate(stats, 2):
        ws2.cell(row=r,column=1,value=label).font = Font(name="Arial",bold=True,size=10)
        ws2.cell(row=r,column=2,value=val).font   = Font(name="Arial",size=10)
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 20

    wb.save(output_path)
    log.info(f"Excel saved → {output_path}")


def send_email(filepath: str, job_count: int, source_summary: dict):
    smtp_host = os.getenv("SMTP_HOST","smtp.gmail.com")
    smtp_port = int(os.getenv("SMTP_PORT",587))
    smtp_user = os.getenv("SMTP_USER")
    smtp_pass = os.getenv("SMTP_PASS")
    to_email  = os.getenv("TO_EMAIL")
    if not all([smtp_user,smtp_pass,to_email]):
        log.warning("Email credentials missing — skipping notification.")
        return
    src_lines = "\n".join(f"  • {src}: {cnt} jobs" for src,cnt in source_summary.items())
    msg = MIMEMultipart()
    msg["From"]    = smtp_user
    msg["To"]      = to_email
    msg["Subject"] = f"DevOps Radar — {job_count} matches — {datetime.now().strftime('%d %b %Y')}"
    msg.attach(MIMEText(
        f"Hi Seetharam,\n\n{job_count} DevOps roles matched your profile today.\n\n"
        f"Sources scanned:\n{src_lines}\n\n"
        f"Scoring: Kubernetes 20% | CI/CD 18% | Cloud 17% | Terraform 15%\n"
        f"Threshold: ≥{MATCH_THRESHOLD}% match\n\n"
        f"Excel attached. Live dashboard → your GitHub Pages URL\n\n— DevOps Radar Bot", "plain"))
    with open(filepath,"rb") as f:
        part = MIMEBase("application","octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition",f"attachment; filename={Path(filepath).name}")
    msg.attach(part)
    try:
        with smtplib.SMTP(smtp_host, smtp_port) as srv:
            srv.starttls()
            srv.login(smtp_user, smtp_pass)
            srv.send_message(msg)
        log.info(f"Email sent → {to_email}")
    except Exception as e:
        log.error(f"Email failed: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def run(docs_dir: str = DOCS_DIR):
    log.info("=" * 65)
    log.info("DevOps Job Radar — Multi-Source Edition")
    log.info("Sources: LinkedIn | Indeed | Adzuna | Arbeitnow | RemoteOK | WorkingNomads | TheMuse")
    log.info("=" * 65)

    raw_jobs  = collect_all_jobs()
    processed = process_jobs(raw_jobs)

    output_dir = Path("output")
    output_dir.mkdir(exist_ok=True)
    filename  = f"devops_jobs_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    filepath  = str(output_dir / filename)

    export_json(processed, docs_dir=docs_dir)

    if processed:
        build_excel(processed, filepath)
        source_summary = {}
        for j in processed:
            s = j.get("source","?")
            source_summary[s] = source_summary.get(s,0) + 1
        send_email(filepath, len(processed), source_summary)

    log.info(f"Done — {len(processed)} jobs written.")
    return filepath, len(processed)


if __name__ == "__main__":
    run()